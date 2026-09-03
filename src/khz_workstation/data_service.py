from __future__ import annotations

from getpass import getuser
import csv
import re
from pathlib import Path

from .schema_sidecar import load_sidecar, sidecar_path, sqlite_types_for_headers, value_violates_schema
from .workspace import Workspace

_SCI = re.compile(r"^[+-]?(?:\d+\.?\d*|\.\d+)[eE][+-]?\d+$")
MAX_SOURCE_BYTES = 32 * 1024 * 1024
MAX_COLUMNS = 256
MAX_ROWS = 5000


class DataDependencyError(RuntimeError):
    pass


def _identifier(value: str, fallback: str) -> str:
    value = re.sub(r"[^A-Za-z0-9_]", "_", value.strip())
    if not value or not value[0].isalpha():
        value = fallback + "_" + value
    return value[:63]


def _dedupe_headers(headers: list[str]) -> list[str]:
    out: list[str] = []
    used: set[str] = set()
    for idx, raw in enumerate(headers, 1):
        base = _identifier(raw or f"Column{idx}", f"Column{idx}")
        name = base
        n = 2
        while name.casefold() in used:
            name = f"{base}_{n}"[:63]
            n += 1
        used.add(name.casefold())
        out.append(name)
    return out


def _looks_leading_zero_int(text: str) -> bool:
    body = text[1:] if text[:1] in "+-" else text
    return body.isdigit() and len(body) > 1 and body.startswith("0")


def _digit_count(text: str) -> int:
    return sum(1 for ch in text if ch.isdigit())


def _infer_type(values: list[object]) -> str:
    nonempty = [v for v in values if v not in (None, "")]
    if not nonempty:
        return "TEXT"
    all_int = True
    all_num = True
    for value in nonempty:
        if isinstance(value, bool):
            return "TEXT"
        if isinstance(value, int):
            if abs(value) >= 10 ** 15:
                return "TEXT"
            continue
        if isinstance(value, float):
            all_int = False
            continue
        text = str(value).strip()
        if _SCI.match(text) or _looks_leading_zero_int(text) or _digit_count(text) > 15:
            return "TEXT"
        try:
            int(text)
        except ValueError:
            all_int = False
        try:
            float(text)
        except ValueError:
            all_num = False
    if all_int:
        return "INTEGER"
    if all_num:
        return "REAL"
    return "TEXT"


def _coerce(value: object, typ: str) -> object | None:
    if value in (None, ""):
        return None
    if typ == "INTEGER":
        return int(str(value).strip(), 10)
    if typ == "REAL":
        return float(value)
    return str(value)


class DataWorkspaceService:
    def __init__(self, workspace: Workspace) -> None:
        self.ws = workspace

    def import_csv(self, source: Path, table_name: str | None = None, *, require_schema: bool = False) -> str:
        source = source.resolve(strict=True)
        self._enforce_source_limits(source)
        with source.open("r", encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh))
        if not rows or not rows[0]:
            raise ValueError("CSV has no header row.")
        return self._import_rows(
            rows[0],
            rows[1:],
            table_name or _identifier(source.stem, "Imported"),
            f"CSV:{source.name}",
            source=source,
            require_schema=require_schema,
        )

    def export_csv(self, table_id: str, destination: Path) -> Path:
        cols, rows = self.ws.store.query_data(table_id, limit=5000)
        destination.parent.mkdir(parents=True, exist_ok=True)
        with destination.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(cols)
            writer.writerows([[row[c] for c in cols] for row in rows])
        self.ws.audit.append(who=getuser(), what="data.export_csv", target=destination.name, approval="USER", execution="DETERMINISTIC", result="EXPORTED", verification=f"rows={len(rows)}")
        return destination

    def import_xlsx(self, source: Path, table_name: str | None = None, sheet_name: str | None = None, *, require_schema: bool = False) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise DataDependencyError("XLSX Data import requires the pinned automation dependency openpyxl.") from exc
        source = source.resolve(strict=True)
        self._enforce_source_limits(source)
        wb = load_workbook(source, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        values = list(ws.iter_rows(values_only=True))
        title = ws.title
        wb.close()
        if not values or not values[0]:
            raise ValueError("Worksheet has no header row.")
        headers = ["" if x is None else str(x) for x in values[0]]
        rows = [list(row) for row in values[1:]]
        return self._import_rows(headers, rows, table_name or _identifier(source.stem, "Imported"), f"XLSX:{source.name}:{title}", source=source, require_schema=require_schema)

    def export_xlsx(self, table_id: str, destination: Path) -> Path:
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise DataDependencyError("XLSX Data export requires the pinned automation dependency openpyxl.") from exc
        cols, rows = self.ws.store.query_data(table_id, limit=5000)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(cols)
        for row in rows:
            ws.append([row[c] for c in cols])
        ws.freeze_panes = "A2"
        destination.parent.mkdir(parents=True, exist_ok=True)
        wb.save(destination)
        self.ws.audit.append(who=getuser(), what="data.export_xlsx", target=destination.name, approval="USER", execution="DETERMINISTIC", result="EXPORTED", verification=f"rows={len(rows)}")
        return destination

    def _enforce_source_limits(self, source: Path) -> None:
        size = source.stat().st_size
        if size > MAX_SOURCE_BYTES:
            raise ValueError(f"LIMIT_SOURCE: {size} bytes exceeds {MAX_SOURCE_BYTES}")

    def _import_rows(self, raw_headers: list[str], rows: list[list[object]], table_name: str, source_label: str, *, source: Path, require_schema: bool) -> str:
        if len(raw_headers) > MAX_COLUMNS:
            raise ValueError(f"LIMIT_COLUMNS: {len(raw_headers)} exceeds {MAX_COLUMNS}")
        if len(rows) > MAX_ROWS:
            raise ValueError(f"LIMIT_ROWS: {len(rows)} exceeds {MAX_ROWS}")
        headers = _dedupe_headers(raw_headers)
        width = len(headers)
        normalized = [list(row[:width]) + [None] * max(0, width - len(row)) for row in rows]
        schema = None
        if sidecar_path(source).is_file() or require_schema:
            schema = load_sidecar(source)
            types = sqlite_types_for_headers(schema, headers)
            name_to_declared = {c.name.casefold(): c.declared_type for c in schema.columns}
            for row in normalized:
                for i, cell in enumerate(row):
                    if cell in (None, ""):
                        continue
                    code = value_violates_schema(str(cell).strip(), name_to_declared[headers[i].casefold()])
                    if code:
                        from .schema_sidecar import SchemaError
                        raise SchemaError(f"{code}: column {headers[i]} value {cell!r}")
        else:
            column_values = [[row[idx] for row in normalized] for idx in range(width)]
            types = [_infer_type(values) for values in column_values]
        converted = [{headers[i]: _coerce(row[i], types[i]) for i in range(width) if row[i] not in (None, "")} for row in normalized]
        table_id = self.ws.store.create_data_table_with_rows(_identifier(table_name, "Imported"), list(zip(headers, types)), converted)
        mode = "SIDECAR" if schema else "INFERRED"
        self.ws.audit.append(who=getuser(), what="data.import", target=table_name, approval="USER", execution="DETERMINISTIC", result="IMPORTED", verification=f"rows={len(normalized)}; contract={mode}; atomic SQLite transaction", metadata={"source": source_label, "columns": headers, "types": types, "contract": mode})
        return table_id
