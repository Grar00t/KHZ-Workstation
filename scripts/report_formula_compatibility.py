from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
SRC=ROOT/"acceptance"/"corpus"/"FormulaCompatibility.xlsx"
RT=ROOT/"acceptance"/"roundtrip"/"FormulaCompatibility.xlsx"
REPORT=ROOT/"acceptance"/"reports"/"formula-compatibility.json"


def main():
    src=load_workbook(SRC,data_only=False)["FormulaTests"]
    rt_f=load_workbook(RT,data_only=False)["FormulaTests"]
    rt_v=load_workbook(RT,data_only=True)["FormulaTests"]
    rows=[]
    for r in range(2,src.max_row+1):
        name=src.cell(r,1).value; before=src.cell(r,2).value; after=rt_f.cell(r,2).value; value=rt_v.cell(r,2).value
        error=isinstance(value,str) and value.startswith("#")
        status="VERIFIED" if after is not None and value is not None and not error else "PARTIAL"
        rows.append({"function":name,"formula_before":before,"formula_after":after,"cached_result":value,"status":status})
    report={"engine":"LibreOffice 25.2.3.2","platform":"linux","rows":rows,"summary":{"VERIFIED":sum(x["status"]=="VERIFIED" for x in rows),"PARTIAL":sum(x["status"]=="PARTIAL" for x in rows)}}
    REPORT.write_text(json.dumps(report,indent=2,sort_keys=True,default=str),encoding="utf-8")
    print(json.dumps(report["summary"],indent=2));
    for x in rows:
        if x["status"]!="VERIFIED": print(x["function"],x["formula_after"],x["cached_result"])
    return 0

if __name__=="__main__": raise SystemExit(main())
