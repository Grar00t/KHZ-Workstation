from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "acceptance" / "corpus" / "FormulaCompatibility.xlsx"

FORMULAS = [
    ("SUM", "=SUM(Data!A2:A6)"),
    ("SUMIF", '=SUMIF(Data!A2:A6,">2",Data!A2:A6)'),
    ("SUMIFS", '=SUMIFS(Data!E2:E6,Data!D2:D6,"OPS")'),
    ("AVERAGE", "=AVERAGE(Data!A2:A6)"),
    ("AVERAGEIF", '=AVERAGEIF(Data!A2:A6,">2",Data!A2:A6)'),
    ("AVERAGEIFS", '=AVERAGEIFS(Data!E2:E6,Data!D2:D6,"OPS")'),
    ("MIN", "=MIN(Data!A2:A6)"),
    ("MAX", "=MAX(Data!A2:A6)"),
    ("COUNT", "=COUNT(Data!A2:A6)"),
    ("COUNTA", "=COUNTA(Data!B2:B6)"),
    ("COUNTIF", '=COUNTIF(Data!D2:D6,"OPS")'),
    ("COUNTIFS", '=COUNTIFS(Data!D2:D6,"OPS",Data!A2:A6,">1")'),
    ("ROUND", "=ROUND(1.2345,2)"),
    ("ROUNDUP", "=ROUNDUP(1.231,2)"),
    ("ROUNDDOWN", "=ROUNDDOWN(1.239,2)"),
    ("SUBTOTAL", "=SUBTOTAL(9,Data!A2:A6)"),
    ("IF", '=IF(Data!A2=1,"yes","no")'),
    ("IFS", '=_xlfn.IFS(Data!A2=1,"one",Data!A2=2,"two",TRUE,"other")'),
    ("AND", "=AND(Data!A2=1,Data!A3=2)"),
    ("OR", "=OR(Data!A2=9,Data!A3=2)"),
    ("NOT", "=NOT(Data!A2=9)"),
    ("IFERROR", '=IFERROR(1/0,"handled")'),
    ("IFNA", '=_xlfn.IFNA(_xlfn.XLOOKUP("MISS",Data!D2:D6,Data!E2:E6),"na")'),
    ("XLOOKUP", '=_xlfn.XLOOKUP("OPS",Data!D2:D6,Data!E2:E6,"missing")'),
    ("VLOOKUP", '=VLOOKUP("OPS",Data!D2:E6,2,FALSE)'),
    ("HLOOKUP", '=HLOOKUP("B",Horizontal!A1:E2,2,FALSE)'),
    ("INDEX", "=INDEX(Data!E2:E6,3)"),
    ("MATCH", '=MATCH("FIN",Data!D2:D6,0)'),
    ("XMATCH", '=_xlfn.XMATCH("FIN",Data!D2:D6,0)'),
    ("OFFSET", "=OFFSET(Data!A2,2,0)"),
    ("INDIRECT", '=INDIRECT("Data!A4")'),
    ("LEFT", '=LEFT("abcdef",2)'),
    ("RIGHT", '=RIGHT("abcdef",2)'),
    ("MID", '=MID("abcdef",2,3)'),
    ("LEN", '=LEN("abcdef")'),
    ("TRIM", '=TRIM("  a   b  ")'),
    ("CONCAT", "=_xlfn.CONCAT(Data!B2:B3)"),
    ("TEXTJOIN", '=_xlfn.TEXTJOIN(",",TRUE,Data!B2:B4)'),
    ("SUBSTITUTE", '=SUBSTITUTE("abcabc","a","x")'),
    ("FIND", '=FIND("cd","abcdef")'),
    ("SEARCH", '=SEARCH("CD","abcdef")'),
    ("TEXT", '=TEXT(1234.5,"0.00")'),
    ("DATE", "=DATE(2026,8,31)"),
    ("TODAY", "=TODAY()"),
    ("NOW", "=NOW()"),
    ("YEAR", "=YEAR(Data!C2)"),
    ("MONTH", "=MONTH(Data!C2)"),
    ("DAY", "=DAY(Data!C2)"),
    ("WORKDAY", "=WORKDAY(Data!C2,5)"),
    ("NETWORKDAYS", "=NETWORKDAYS(Data!C2,Data!C6)"),
    ("FILTER", "=_xlfn._xlws.FILTER(Data!A2:A6,Data!A2:A6>2)"),
    ("SORT", "=_xlfn._xlws.SORT(Data!B2:B6)"),
    ("SORTBY", "=_xlfn._xlws.SORTBY(Data!B2:B6,Data!A2:A6,-1)"),
    ("UNIQUE", "=_xlfn.UNIQUE(Data!D2:D6)"),
    ("SEQUENCE", "=_xlfn.SEQUENCE(3,1,10,2)"),
    ("TAKE", "=_xlfn.TAKE(Data!A2:A6,2)"),
    ("DROP", "=_xlfn.DROP(Data!A2:A6,2)"),
    ("HSTACK", "=_xlfn.HSTACK(Data!A2:A3,Data!B2:B3)"),
    ("VSTACK", "=_xlfn.VSTACK(Data!A2:A3,Data!A4:A5)"),
    ("RELATIVE_REF", "=Data!A2+Data!A3"),
    ("ABSOLUTE_REF", "=Data!$A$2+Data!$A$3"),
    ("MIXED_REF", "=Data!$A2+Data!A$3"),
    ("CROSS_SHEET_REF", "=Data!E2"),
    ("NAMED_RANGE", "=SUM(Values)"),
]


def main() -> int:
    wb = Workbook()
    ws = wb.active; ws.title = "FormulaTests"
    data = wb.create_sheet("Data")
    hor = wb.create_sheet("Horizontal")
    ws.append(["Function", "Formula", "Notes"])
    data.append(["Number", "Text", "Date", "Department", "Amount"])
    values = [
        (1, "alpha", date(2026, 8, 27), "OPS", 100.0),
        (2, "beta", date(2026, 8, 28), "FIN", 200.0),
        (3, "gamma", date(2026, 8, 29), "OPS", 300.0),
        (4, "delta", date(2026, 8, 30), "HR", 400.0),
        (5, "epsilon", date(2026, 8, 31), "IT", 500.0),
    ]
    for row in values: data.append(row)
    hor.append(["A", "B", "C", "D", "E"]); hor.append([10,20,30,40,50])
    wb.defined_names.add(DefinedName("Values", attr_text="Data!$A$2:$A$6"))
    for name, formula in FORMULAS:
        ws.append([name, formula, "SYNTHETIC TEST DATA"])
    ws.freeze_panes = "A2"
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
