from __future__ import annotations

import json
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CORPUS = ROOT / "acceptance" / "corpus"
ROUNDTRIP = ROOT / "acceptance" / "roundtrip"
REPORT = ROOT / "acceptance" / "reports" / "compatibility-structure.json"
NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
}


def zip_features(path: Path) -> dict:
    with zipfile.ZipFile(path) as zf:
        names = set(zf.namelist())
        out = {
            "entries": len(names),
            "pivot_tables": len([n for n in names if n.startswith("xl/pivotTables/") and n.endswith(".xml")]),
            "pivot_cache": len([n for n in names if n.startswith("xl/pivotCache/") and n.endswith(".xml")]),
            "charts": len([n for n in names if "/charts/chart" in n and n.endswith(".xml")]),
            "comments": len([n for n in names if "comments" in n and n.endswith(".xml")]),
            "headers": len([n for n in names if n.startswith("word/header")]),
            "footers": len([n for n in names if n.startswith("word/footer")]),
            "notes_slides": len([n for n in names if n.startswith("ppt/notesSlides/notesSlide") and n.endswith(".xml")]),
            "slide_masters": len([n for n in names if n.startswith("ppt/slideMasters/slideMaster") and n.endswith(".xml")]),
        }
        if "word/document.xml" in names:
            root = ET.fromstring(zf.read("word/document.xml"))
            out["tracked_insertions"] = len(root.findall(".//w:ins", NS))
            out["toc_fields"] = len([x for x in root.iter() if x.attrib.get("{" + NS["w"] + "}instr", "").startswith("TOC")])
        slide_xmls = [n for n in names if n.startswith("ppt/slides/slide") and n.endswith(".xml")]
        out["transitions"] = 0; out["animations"] = 0
        for name in slide_xmls:
            root = ET.fromstring(zf.read(name))
            out["transitions"] += len(root.findall(".//p:transition", NS))
            out["animations"] += len(root.findall(".//p:timing", NS))
        return out


def xlsx_features(path: Path) -> dict:
    wb = load_workbook(path, data_only=False, read_only=False)
    formulas = 0; comments = 0; validations = 0; cf_ranges = 0; charts = 0; tables = 0; protected = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="): formulas += 1
                if cell.comment: comments += 1
        validations += len(ws.data_validations.dataValidation)
        cf_ranges += len(ws.conditional_formatting)
        charts += len(ws._charts)
        tables += len(ws.tables)
        protected += int(bool(ws.protection.sheet))
    z = zip_features(path)
    return {
        "worksheets": len(wb.sheetnames), "sheet_names": wb.sheetnames,
        "formulas": formulas, "defined_names": len(wb.defined_names), "tables": tables,
        "validations": validations, "conditional_formatting_ranges": cf_ranges,
        "charts_openpyxl": charts, "comments_openpyxl": comments, "protected_sheets": protected,
        "pivot_tables_xml": z["pivot_tables"], "pivot_cache_xml": z["pivot_cache"],
    }


def compare(before: dict, after: dict) -> dict:
    keys = sorted(set(before) | set(after))
    return {k: {"before": before.get(k), "after": after.get(k), "same": before.get(k) == after.get(k)} for k in keys}


def main() -> int:
    report = {"xlsx": {}, "docx": {}, "pptx": {}}
    src_x = CORPUS / "InstitutionalWorkbook.xlsx"; rt_x = ROUNDTRIP / src_x.name
    report["xlsx"]["before"] = xlsx_features(src_x)
    if rt_x.exists():
        report["xlsx"]["after"] = xlsx_features(rt_x)
        report["xlsx"]["compare"] = compare(report["xlsx"]["before"], report["xlsx"]["after"])
    src_d = CORPUS / "InstitutionalReport.docx"; rt_d = ROUNDTRIP / src_d.name
    report["docx"]["before"] = zip_features(src_d)
    if rt_d.exists(): report["docx"]["after"] = zip_features(rt_d); report["docx"]["compare"] = compare(report["docx"]["before"], report["docx"]["after"])
    src_p = CORPUS / "InstitutionalPresentation.pptx"; rt_p = ROUNDTRIP / src_p.name
    report["pptx"]["before"] = zip_features(src_p)
    if rt_p.exists(): report["pptx"]["after"] = zip_features(rt_p); report["pptx"]["compare"] = compare(report["pptx"]["before"], report["pptx"]["after"])
    REPORT.parent.mkdir(parents=True, exist_ok=True); REPORT.write_text(json.dumps(report, indent=2, sort_keys=True), encoding="utf-8")
    print(json.dumps(report, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
